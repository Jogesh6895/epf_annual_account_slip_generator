"""
EPF Annual Account Slip Generator
A comprehensive tool for calculating Employees' Provident Fund annual accounts.

Features:
- Excel input/output support with validation
- Comprehensive error handling
- EPF contribution calculations (12% employee, 8.33% EPS, 3.67% employer)
- Interest calculations with monthly balance tracking
- Withdrawal processing
"""

import csv
import subprocess
import time
from typing import List, Dict, Tuple, Union, Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, colors


class EPFCalculatorError(Exception):
    """Base exception for EPF Calculator errors."""

    pass


class FileLoadError(EPFCalculatorError):
    """Exception raised when file cannot be loaded."""

    pass


class SheetNotFoundError(EPFCalculatorError):
    """Exception raised when required sheet is missing."""

    pass


class DataValidationError(EPFCalculatorError):
    """Exception raised when data validation fails."""

    pass


# Constants
EMPLOYEE_CONTRIBUTION_RATE = 0.12
EPS_CONTRIBUTION_RATE = 0.0833
EMPLOYER_CONTRIBUTION_RATE = 0.0367
INTEREST_DIVISOR = 1200
EXPECTED_WAGES_COLUMNS = 14
EXPECTED_WDL_COLUMNS = 12
EXPECTED_OB_COLUMNS = 1


def clear_screen_with_title(title: str) -> None:
    """Clear console screen and set title (Windows only)."""
    subprocess.call("pause >nul", shell=True)
    subprocess.call("cls", shell=True)
    subprocess.call(f"title={title}", shell=True)
    print(title + "\n")


def start_up_check() -> str:
    """Display startup instructions and get user confirmation."""
    clear_screen_with_title("EPF Annual Account Slip Generator")
    subprocess.call("color 0A", shell=True)
    print("Instructions:\n")
    print(
        "1. Make sure that you have filled all the sheets present in the 'Input.xlsx' "
        "file properly before running the program."
    )
    print("2. Existing 'Output.xlsx' and 'output.csv' files will be overwritten.\n")
    while True:
        response = input("Enter 'y' to Continue or 'n' to Quit: ")
        if response in ("y", "Y", "n", "N"):
            return response
        print("Invalid Choice! Please Enter Either 'y' to Continue or 'n' to Quit.\n")


def open_input_excel_file(path: str) -> openpyxl.Workbook:
    """Load Excel workbook from specified path.

    Args:
        path: Path to Excel file

    Returns:
        openpyxl.Workbook: Loaded workbook

    Raises:
        FileLoadError: If file cannot be loaded
    """
    try:
        return openpyxl.load_workbook(path)
    except FileNotFoundError:
        raise FileLoadError(f"File not found: {path}")
    except Exception as e:
        raise FileLoadError(f"Failed to load workbook: {e}")


def get_sheet(
    workbook: openpyxl.Workbook, sheet_name: str
) -> openpyxl.worksheet.worksheet.Worksheet:
    """Get worksheet from workbook.

    Args:
        workbook: Excel workbook
        sheet_name: Name of sheet to retrieve

    Returns:
        Worksheet object

    Raises:
        SheetNotFoundError: If sheet doesn't exist
    """
    try:
        return workbook[sheet_name]
    except KeyError:
        raise SheetNotFoundError(f"Sheet '{sheet_name}' not found in workbook")


def validate_sheet_dimensions(
    wage_sheet: openpyxl.worksheet.worksheet.Worksheet,
    ob_ee_sheet: openpyxl.worksheet.worksheet.Worksheet,
    ob_er_sheet: openpyxl.worksheet.worksheet.Worksheet,
    ob_eps_sheet: openpyxl.worksheet.worksheet.Worksheet,
    wdl_ee_sheet: openpyxl.worksheet.worksheet.Worksheet,
    wdl_er_sheet: openpyxl.worksheet.worksheet.Worksheet,
) -> None:
    """Validate column and row counts for all sheets.

    Args:
        wage_sheet: Wages sheet
        ob_ee_sheet: OB_EE sheet
        ob_er_sheet: OB_ER sheet
        ob_eps_sheet: OB_EPS sheet
        wdl_ee_sheet: WDL_EE sheet
        wdl_er_sheet: WDL_ER sheet

    Raises:
        DataValidationError: If validation fails
    """
    wage_sheet_row_count = wage_sheet.max_row

    # Column validation
    if wage_sheet.max_column != EXPECTED_WAGES_COLUMNS:
        raise DataValidationError(
            f"In 'Wages' Sheet Expected {EXPECTED_WAGES_COLUMNS} columns, "
            f"found {wage_sheet.max_column}."
        )
    if ob_ee_sheet.max_column != EXPECTED_OB_COLUMNS:
        raise DataValidationError(
            f"In 'OB_EE' Sheet Expected {EXPECTED_OB_COLUMNS} column, "
            f"found {ob_ee_sheet.max_column}."
        )
    if ob_er_sheet.max_column != EXPECTED_OB_COLUMNS:
        raise DataValidationError(
            f"In 'OB_ER' Sheet Expected {EXPECTED_OB_COLUMNS} column, "
            f"found {ob_er_sheet.max_column}."
        )
    if ob_eps_sheet.max_column != EXPECTED_OB_COLUMNS:
        raise DataValidationError(
            f"In 'OB_EPS' Sheet Expected {EXPECTED_OB_COLUMNS} column, "
            f"found {ob_eps_sheet.max_column}."
        )
    if wdl_ee_sheet.max_column != EXPECTED_WDL_COLUMNS:
        raise DataValidationError(
            f"In 'WDL_EE' Sheet Expected {EXPECTED_WDL_COLUMNS} columns, "
            f"found {wdl_ee_sheet.max_column}."
        )
    if wdl_er_sheet.max_column != EXPECTED_WDL_COLUMNS:
        raise DataValidationError(
            f"In 'WDL_ER' Sheet Expected {EXPECTED_WDL_COLUMNS} columns, "
            f"found {wdl_er_sheet.max_column}."
        )

    # Row validation
    if wage_sheet_row_count != ob_ee_sheet.max_row:
        raise DataValidationError(
            f"Row count mismatch: 'Wages' has {wage_sheet_row_count} rows, "
            f"'OB_EE' has {ob_ee_sheet.max_row} rows."
        )
    if wage_sheet_row_count != ob_er_sheet.max_row:
        raise DataValidationError(
            f"Row count mismatch: 'Wages' has {wage_sheet_row_count} rows, "
            f"'OB_ER' has {ob_er_sheet.max_row} rows."
        )
    if wage_sheet_row_count != ob_eps_sheet.max_row:
        raise DataValidationError(
            f"Row count mismatch: 'Wages' has {wage_sheet_row_count} rows, "
            f"'OB_EPS' has {ob_eps_sheet.max_row} rows."
        )
    if wage_sheet_row_count != wdl_ee_sheet.max_row:
        raise DataValidationError(
            f"Row count mismatch: 'Wages' has {wage_sheet_row_count} rows, "
            f"'WDL_EE' has {wdl_ee_sheet.max_row} rows."
        )
    if wage_sheet_row_count != wdl_er_sheet.max_row:
        raise DataValidationError(
            f"Row count mismatch: 'Wages' has {wage_sheet_row_count} rows, "
            f"'WDL_ER' has {wdl_er_sheet.max_row} rows."
        )


def fetch_sheet_data(
    sheet: openpyxl.worksheet.worksheet.Worksheet, data_type: str
) -> Union[List, List[List]]:
    """Fetch data from sheet starting from row 2.

    Args:
        sheet: Worksheet to fetch data from
        data_type: Type of data ('single' or 'multi')

    Returns:
        List of values or List of lists
    """
    data = []
    for row in range(2, sheet.max_row + 1):
        if data_type == "single":
            value = sheet.cell(row, 1).value
            data.append(0 if value is None else value)
        else:
            row_data = []
            for column in range(1, sheet.max_column + 1):
                value = sheet.cell(row, column).value
                row_data.append(0 if value is None else value)
            data.append(row_data)
    return data


def calculate_contributions(wage: float) -> Tuple[int, int, int]:
    """Calculate EPF contributions from wage.

    Args:
        wage: Monthly wage amount

    Returns:
        Tuple of (employee_contribution, employer_contribution, eps_contribution)
    """
    employee = round(wage * EMPLOYEE_CONTRIBUTION_RATE)
    eps = round(wage * EPS_CONTRIBUTION_RATE)
    employer = round(wage * EMPLOYER_CONTRIBUTION_RATE)
    return employee, employer, eps


def calculate_monthly_balances(
    ob_ee: int,
    ob_er: int,
    ee_contributions: List[int],
    er_contributions: List[int],
    ee_withdrawals: List[int],
    er_withdrawals: List[int],
    rate: float,
) -> Tuple[List[int], List[int], int, int]:
    """Calculate monthly balances and interest for employee and employer shares.

    Args:
        ob_ee: Opening balance for employee share
        ob_er: Opening balance for employer share
        ee_contributions: Monthly employee contributions
        er_contributions: Monthly employer contributions
        ee_withdrawals: Monthly employee withdrawals
        er_withdrawals: Monthly employer withdrawals
        rate: Annual interest rate

    Returns:
        Tuple of (ee_balances, er_balances, ee_interest, er_interest)
    """
    ee_balances = [ob_ee]
    er_balances = [ob_er]
    current_ee_balance = ob_ee
    current_er_balance = ob_er

    for i in range(len(ee_contributions) - 1):
        current_ee_balance = (
            current_ee_balance + ee_contributions[i] - ee_withdrawals[i]
        )
        current_er_balance = (
            current_er_balance + er_contributions[i] - er_withdrawals[i]
        )
        ee_balances.append(current_ee_balance)
        er_balances.append(current_er_balance)

    sum_ee_balances = sum(ee_balances)
    sum_er_balances = sum(er_balances)
    total_ee_withdrawals = sum(ee_withdrawals)
    total_er_withdrawals = sum(er_withdrawals)

    ee_interest = round(
        ((sum_ee_balances - total_ee_withdrawals) * rate) / INTEREST_DIVISOR
    )
    er_interest = round(
        ((sum_er_balances - total_er_withdrawals) * rate) / INTEREST_DIVISOR
    )

    return ee_balances, er_balances, ee_interest, er_interest


def generate_output_rows(
    wage_sheet_data: List[List],
    ob_ee_data: List,
    ob_er_data: List,
    ob_eps_data: List,
    wdl_ee_data: List[List],
    wdl_er_data: List[List],
    rate: float,
) -> List[List]:
    """Generate output rows with all calculations.

    Args:
        wage_sheet_data: Wage data (rows of monthly wages)
        ob_ee_data: Opening balances for employee share
        ob_er_data: Opening balances for employer share
        ob_eps_data: Opening balances for EPS
        wdl_ee_data: Monthly withdrawals from employee share
        wdl_er_data: Monthly withdrawals from employer share
        rate: Annual interest rate

    Returns:
        List of output rows
    """
    output_rows = [
        [
            "A/C No.",
            "NAME",
            "OB(EE)",
            "OB(ER)",
            "INT(EE)",
            "INT(ER)",
            "CONT(EE)",
            "CONT(ER)",
            "WDL(EE)",
            "WDL(ER)",
            "CB(EE)",
            "CB(ER)",
            "OB(EPS)",
            "CONT(EPS)",
            "CB(EPS)",
        ]
    ]

    for i, wage_row in enumerate(wage_sheet_data):
        account_no = wage_row[0]
        name = wage_row[1]
        monthly_wages = wage_row[2 : 2 + EXPECTED_WDL_COLUMNS]

        ee_contribs = []
        er_contribs = []
        eps_contribs = []

        for wage in monthly_wages:
            ee, er, eps = calculate_contributions(wage)
            ee_contribs.append(ee)
            er_contribs.append(er)
            eps_contribs.append(eps)

        ob_ee = int(ob_ee_data[i])
        ob_er = int(ob_er_data[i])
        ob_eps = int(ob_eps_data[i])

        ee_withdrawals = [int(x) for x in wdl_ee_data[i]]
        er_withdrawals = [int(x) for x in wdl_er_data[i]]

        ee_balances, er_balances, ee_interest, er_interest = calculate_monthly_balances(
            ob_ee, ob_er, ee_contribs, er_contribs, ee_withdrawals, er_withdrawals, rate
        )

        total_ee_contrib = sum(ee_contribs)
        total_er_contrib = sum(er_contribs)
        total_ee_wdl = sum(ee_withdrawals)
        total_er_wdl = sum(er_withdrawals)
        total_eps = sum(eps_contribs)

        cb_ee = ob_ee + ee_interest + total_ee_contrib - total_ee_wdl
        cb_er = ob_er + er_interest + total_er_contrib - total_er_wdl
        cb_eps = ob_eps + total_eps

        output_rows.append(
            [
                account_no,
                name,
                str(ob_ee),
                str(ob_er),
                str(ee_interest),
                str(er_interest),
                str(total_ee_contrib),
                str(total_er_contrib),
                str(total_ee_wdl),
                str(total_er_wdl),
                str(cb_ee),
                str(cb_er),
                str(ob_eps),
                str(total_eps),
                str(cb_eps),
            ]
        )

    return output_rows


def write_csv_output(output_rows: List[List], filename: str = "output.csv") -> None:
    """Write output data to CSV file.

    Args:
        output_rows: List of rows to write
        filename: Output filename
    """
    with open(filename, "w", newline="") as csvfile:
        writer = csv.writer(csvfile, delimiter=",")
        for row in output_rows:
            writer.writerow(row)


def write_excel_output(output_rows: List[List], filename: str = "Output.xlsx") -> None:
    """Write output data to Excel file with formatting.

    Args:
        output_rows: List of rows to write
        filename: Output filename
    """
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "EPF Annual Account Slip"

    for row_idx, row in enumerate(output_rows, start=1):
        for col_idx, value in enumerate(row, start=1):
            cell = sheet.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.font = Font(size=12, color=colors.RED, italic=True, bold=True)
                cell.fill = PatternFill("solid", fgColor="7FFFD4")
                cell.border = Border(
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                    top=Side(style="thin"),
                    bottom=Side(style="double"),
                )
                cell.alignment = Alignment(horizontal="center", vertical="center")

    workbook.save(filename)
    workbook.close()


def main() -> Optional[str]:
    """Main function to run EPF calculator."""
    if start_up_check() in ("n", "N"):
        return "CLOSE"

    clear_screen_with_title("Loading Data...")
    path_of_file = "InputFiles/Input.xlsx"

    try:
        input_workbook = open_input_excel_file(path_of_file)
        print("Successfully Loaded the 'Input.xlsx' File.\n")

        wage_sheet = get_sheet(input_workbook, "Wages")
        print("Successfully Loaded the 'Wages' Sheet.")

        ob_ee_sheet = get_sheet(input_workbook, "OB_EE")
        print("Successfully Loaded the 'OB_EE' Sheet.")

        ob_er_sheet = get_sheet(input_workbook, "OB_ER")
        print("Successfully Loaded the 'OB_ER' Sheet.")

        ob_eps_sheet = get_sheet(input_workbook, "OB_EPS")
        print("Successfully Loaded the 'OB_EPS' Sheet.")

        wdl_ee_sheet = get_sheet(input_workbook, "WDL_EE")
        print("Successfully Loaded the 'WDL_EE' Sheet.")

        wdl_er_sheet = get_sheet(input_workbook, "WDL_ER")
        print("Successfully Loaded the 'WDL_ER' Sheet.")

        print("\nAll Sheets Loaded Successfully!")

        clear_screen_with_title("Validating Data...")
        validate_sheet_dimensions(
            wage_sheet,
            ob_ee_sheet,
            ob_er_sheet,
            ob_eps_sheet,
            wdl_ee_sheet,
            wdl_er_sheet,
        )
        print("Data Validation Successful!\n")

        clear_screen_with_title("Processing Data...")
        wage_sheet_data = fetch_sheet_data(wage_sheet, "multi")
        ob_ee_data = fetch_sheet_data(ob_ee_sheet, "single")
        ob_er_data = fetch_sheet_data(ob_er_sheet, "single")
        ob_eps_data = fetch_sheet_data(ob_eps_sheet, "single")
        wdl_ee_data = fetch_sheet_data(wdl_ee_sheet, "multi")
        wdl_er_data = fetch_sheet_data(wdl_er_sheet, "multi")

        print("Data Fetched Successfully.\n")
        rate = float(input("Enter the Rate of Interest for the Year: "))

        print("\nCalculating EPF Contributions and Interest...")
        output_rows = generate_output_rows(
            wage_sheet_data,
            ob_ee_data,
            ob_er_data,
            ob_eps_data,
            wdl_ee_data,
            wdl_er_data,
            rate,
        )
        print("Calculations Complete.\n")

        clear_screen_with_title("Generating Output Files...")
        write_csv_output(output_rows)
        print("Successfully Generated 'output.csv'")

        write_excel_output(output_rows)
        print("Successfully Generated 'Output.xlsx'")

        input_workbook.close()

        return "SUCCESS"

    except EPFCalculatorError as e:
        subprocess.call("color 0C", shell=True)
        print(f"\n####\nERROR: {e}")
        return "ERROR"
    except Exception as e:
        subprocess.call("color 0C", shell=True)
        print(f"\n####\nUNEXPECTED ERROR: {e}")
        return "ERROR"


if __name__ == "__main__":
    starttime = time.time()

    result = main()

    if result == "CLOSE":
        subprocess.call("exit", shell=True)
    elif result == "SUCCESS":
        subprocess.call("title=!!!Program Completed!!!", shell=True)
        subprocess.call("color D0", shell=True)
        print(f"\nTotal Time Taken: {time.time() - starttime:.2f} Seconds")
        print("!!!Program Completed!!!")
        subprocess.call("pause >nul", shell=True)
    else:
        subprocess.call("pause >nul", shell=True)
