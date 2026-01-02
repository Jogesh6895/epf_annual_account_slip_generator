"""
Create sample input file for EPF Report Generator.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, colors


def create_sample_input_file():
    """Create sample Input.xlsx with dummy data."""

    workbook = openpyxl.Workbook()

    # Remove default sheet
    workbook.remove(workbook.active)

    # Create Wages sheet
    wages_sheet = workbook.create_sheet("Wages")
    headers_wages = ["A/C No.", "NAME"] + [
        "Apr",
        "May",
        "Jun",
        "Jul",
        "Aug",
        "Sep",
        "Oct",
        "Nov",
        "Dec",
        "Jan",
        "Feb",
        "Mar",
    ]

    for col_idx, header in enumerate(headers_wages, start=1):
        cell = wages_sheet.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill("solid", fgColor="D3D3D3")
        cell.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Add sample employee data
    sample_employees = [
        [
            "EPF001",
            "John Doe",
            15000,
            15000,
            15500,
            15000,
            15000,
            15000,
            16000,
            15000,
            15000,
            15500,
            15000,
            15000,
        ],
        [
            "EPF002",
            "Jane Smith",
            18000,
            18000,
            18500,
            18000,
            18000,
            18000,
            19000,
            18000,
            18000,
            18500,
            18000,
            18000,
        ],
        [
            "EPF003",
            "Raj Kumar",
            12000,
            12000,
            12500,
            12000,
            12000,
            12000,
            13000,
            12000,
            12000,
            12500,
            12000,
            12000,
        ],
        [
            "EPF004",
            "Sunita Sharma",
            20000,
            20000,
            20500,
            20000,
            20000,
            20000,
            21000,
            20000,
            20000,
            20500,
            20000,
            20000,
        ],
        [
            "EPF005",
            "Amit Patel",
            25000,
            25000,
            25500,
            25000,
            25000,
            25000,
            26000,
            25000,
            25000,
            25500,
            25000,
            25000,
        ],
    ]

    for row_idx, emp_data in enumerate(sample_employees, start=2):
        for col_idx, value in enumerate(emp_data, start=1):
            cell = wages_sheet.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # Create OB_EE sheet
    ob_ee_sheet = workbook.create_sheet("OB_EE")
    cell = ob_ee_sheet.cell(row=1, column=1, value="OB(EE)")
    cell.font = Font(bold=True, size=11)
    cell.fill = PatternFill("solid", fgColor="D3D3D3")
    cell.border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    cell.alignment = Alignment(horizontal="center", vertical="center")

    ob_ee_values = [50000, 60000, 40000, 65000, 75000]
    for row_idx, value in enumerate(ob_ee_values, start=2):
        cell = ob_ee_sheet.cell(row=row_idx, column=1, value=value)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Create OB_ER sheet
    ob_er_sheet = workbook.create_sheet("OB_ER")
    cell = ob_er_sheet.cell(row=1, column=1, value="OB(ER)")
    cell.font = Font(bold=True, size=11)
    cell.fill = PatternFill("solid", fgColor="D3D3D3")
    cell.border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    cell.alignment = Alignment(horizontal="center", vertical="center")

    ob_er_values = [15000, 18000, 12000, 19500, 22500]
    for row_idx, value in enumerate(ob_er_values, start=2):
        cell = ob_er_sheet.cell(row=row_idx, column=1, value=value)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Create OB_EPS sheet
    ob_eps_sheet = workbook.create_sheet("OB_EPS")
    cell = ob_eps_sheet.cell(row=1, column=1, value="OB(EPS)")
    cell.font = Font(bold=True, size=11)
    cell.fill = PatternFill("solid", fgColor="D3D3D3")
    cell.border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    cell.alignment = Alignment(horizontal="center", vertical="center")

    ob_eps_values = [35000, 42000, 28000, 45500, 52500]
    for row_idx, value in enumerate(ob_eps_values, start=2):
        cell = ob_eps_sheet.cell(row=row_idx, column=1, value=value)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Create WDL_EE sheet
    wdl_ee_sheet = workbook.create_sheet("WDL_EE")
    wdl_headers = [
        "Apr",
        "May",
        "Jun",
        "Jul",
        "Aug",
        "Sep",
        "Oct",
        "Nov",
        "Dec",
        "Jan",
        "Feb",
        "Mar",
    ]

    for col_idx, header in enumerate(wdl_headers, start=1):
        cell = wdl_ee_sheet.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill("solid", fgColor="D3D3D3")
        cell.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        cell.alignment = Alignment(horizontal="center", vertical="center")

    wdl_ee_data = [
        [0, 0, 0, 5000, 0, 0, 0, 0, 0, 0, 0, 0],
        [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
        [0, 0, 0, 0, 0, 10000, 0, 0, 0, 0, 0, 0],
        [0, 8000, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
        [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
    ]

    for row_idx, row_data in enumerate(wdl_ee_data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            cell = wdl_ee_sheet.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # Create WDL_ER sheet
    wdl_er_sheet = workbook.create_sheet("WDL_ER")

    for col_idx, header in enumerate(wdl_headers, start=1):
        cell = wdl_er_sheet.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill("solid", fgColor="D3D3D3")
        cell.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        cell.alignment = Alignment(horizontal="center", vertical="center")

    wdl_er_data = [
        [0, 0, 0, 1500, 0, 0, 0, 0, 0, 0, 0, 0],
        [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
        [0, 0, 0, 0, 0, 3300, 0, 0, 0, 0, 0, 0],
        [0, 2670, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
        [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
    ]

    for row_idx, row_data in enumerate(wdl_er_data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            cell = wdl_er_sheet.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # Save workbook
    workbook.save("InputFiles/Sample_Input.xlsx")
    workbook.close()
    print("Sample Input.xlsx file created successfully in InputFiles directory.")
    print("\nSample data includes:")
    print(f"- 5 sample employees")
    print(f"- 12 months of wage data")
    print(f"- Opening balances for EE, ER, and EPS")
    print(f"- Withdrawal data (some employees have withdrawals)")
    print(f"\nFile: InputFiles/Sample_Input.xlsx")


if __name__ == "__main__":
    create_sample_input_file()
